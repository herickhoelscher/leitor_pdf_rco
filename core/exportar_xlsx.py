"""
Exporta um arquivo Excel (.xlsx) com 3 abas:
  1. Frequência Detalhada  — uma linha por aluno por aula (ideal para Power BI)
  2. Resumo por Aluno       — totais consolidados por aluno/turma/disciplina
  3. Conteúdos              — uma linha por aula com conteúdo/observações

Campos calculados adicionados:
  - total_aulas, taxa_presenca_pct, status_frequencia, data_formatada
"""

import re
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────
_AZUL_ESC  = "1E3A6E"
_AZUL_MED  = "2563EB"
_CINZA_CLR = "F1F5F9"
_CINZA_BRD = "CBD5E1"
_VERDE     = "16A34A"
_AMARELO   = "D97706"
_VERMELHO  = "DC2626"
_BRANCO    = "FFFFFF"

MESES = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

LIMITE_RISCO    = 0.75   # abaixo → Reprovado (frequência mínima obrigatória)
LIMITE_ATENCAO  = 0.80   # abaixo → Em Risco (margem de segurança)


def _normalizar_data(data_str, ano=None):
    """Converte '16 set' → '16/09' ou '16/09/2025' se ano fornecido."""
    partes = data_str.strip().split()
    if len(partes) >= 2:
        dia = partes[0].zfill(2)
        mes = MESES.get(partes[1].lower())
        if mes:
            if ano:
                return f"{dia}/{mes:02d}/{ano}"
            return f"{dia}/{mes:02d}"
    return data_str


def _status(taxa_presenca):
    """Retorna status de frequência com base na taxa (0.0–1.0)."""
    if taxa_presenca >= LIMITE_ATENCAO:
        return "Regular"
    if taxa_presenca >= LIMITE_RISCO:
        return "Em Risco"
    return "Reprovado"


def _status_cor(status):
    if status == "Regular":
        return _VERDE
    if status == "Em Risco":
        return _AMARELO
    return _VERMELHO


def _presenca_cor(marca):
    if marca == "F":
        return _VERMELHO
    if marca == "D":
        return _AMARELO
    return None  # C = sem destaque


# ── Helpers de estilo ─────────────────────────────────────────────
def _font(bold=False, color=None, size=11, italic=False):
    return Font(bold=bold, color=color or "000000", size=size, italic=italic)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _border_thin():
    s = Side(style="thin", color=_CINZA_BRD)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_row(ws, colunas, row=1, cor_bg=_AZUL_ESC):
    for col, titulo in enumerate(colunas, 1):
        c = ws.cell(row=row, column=col, value=titulo)
        c.font = _font(bold=True, color=_BRANCO, size=10)
        c.fill = _fill(cor_bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border_thin()


def _auto_width(ws, extra=4):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + extra, 60)


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ── Aba 1: Frequência Detalhada ───────────────────────────────────
def _aba_frequencia(wb, dados):
    ws = wb.active
    ws.title = "Frequência Detalhada"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30

    colunas = [
        "Arquivo", "Escola", "Turma", "Disciplina", "Trimestre",
        "Série", "Período", "Aluno", "Data", "Data Formatada",
        "Presença", "Total Aulas", "Total Faltas (calc.)",
        "Taxa Presença (%)", "Status",
        "Tipo de Aula", "Conteúdo / Observações", "Responsável",
    ]
    _header_row(ws, colunas)
    _freeze(ws)

    for i, d in enumerate(dados, 2):
        taxa = d.get("taxa_presenca", 0.0)
        st   = _status(taxa)

        vals = [
            d.get("arquivo", ""),
            d.get("escola", ""),
            d.get("turma", ""),
            d.get("disciplina", ""),
            d.get("trimestre", ""),
            d.get("serie", ""),
            d.get("periodo", ""),
            d.get("aluno", ""),
            d.get("data", ""),
            d.get("data_fmt", ""),
            d.get("presenca", ""),
            d.get("total_aulas", ""),
            d.get("total_faltas_calc", ""),
            round(taxa * 100, 1) if isinstance(taxa, float) else "",
            st,
            d.get("tipo_aula", ""),
            d.get("conteudo_obs", ""),
            d.get("responsavel", ""),
        ]

        alt = (i % 2 == 0)
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.border = _border_thin()
            c.alignment = Alignment(vertical="center", wrap_text=(col == 17))
            c.font = _font(size=10)

            # Zebra
            if alt:
                c.fill = _fill(_CINZA_CLR)

            # Destaque de presença (coluna 11)
            if col == 11:
                cor = _presenca_cor(val)
                if cor:
                    c.font = _font(bold=True, color=cor, size=10)

            # Destaque de status (coluna 15)
            if col == 15:
                c.font = _font(bold=True, color=_status_cor(st), size=10)

            # Taxa como percentual (coluna 14)
            if col == 14 and isinstance(val, float):
                c.number_format = "0.0%"
                c.value = taxa  # armazena como decimal, exibe como %

    _auto_width(ws)
    ws.column_dimensions["Q"].width = 50  # conteúdo mais largo


# ── Aba 2: Resumo por Aluno ───────────────────────────────────────
def _aba_resumo(wb, dados):
    ws = wb.create_sheet("Resumo por Aluno")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30

    colunas = [
        "Escola", "Turma", "Disciplina", "Trimestre", "Série",
        "Período", "Aluno", "Total Aulas", "Presenças", "Faltas",
        "Dispensados", "Taxa Presença (%)", "Status",
    ]
    _header_row(ws, colunas, cor_bg=_AZUL_MED)
    _freeze(ws)

    # Agrupa por (escola, turma, disciplina, trimestre, aluno)
    resumo = {}
    for d in dados:
        chave = (
            d.get("escola", ""), d.get("turma", ""), d.get("disciplina", ""),
            d.get("trimestre", ""), d.get("serie", ""), d.get("periodo", ""),
            d.get("aluno", ""),
        )
        if chave not in resumo:
            resumo[chave] = {"C": 0, "F": 0, "D": 0, "total": 0}
        p = d.get("presenca", "")
        if p in ("C", "F", "D"):
            resumo[chave][p] += 1
        resumo[chave]["total"] = d.get("total_aulas", 0) or resumo[chave]["total"]

    for i, (chave, cont) in enumerate(sorted(resumo.items()), 2):
        escola, turma, disc, trim, serie, periodo, aluno = chave
        total   = cont["total"] or (cont["C"] + cont["F"] + cont["D"])
        faltas  = cont["F"]
        presenc = cont["C"]
        disp    = cont["D"]
        taxa    = presenc / total if total > 0 else 1.0
        st      = _status(taxa)
        alt     = (i % 2 == 0)

        vals = [escola, turma, disc, trim, serie, periodo, aluno,
                total, presenc, faltas, disp, taxa, st]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.border = _border_thin()
            c.alignment = Alignment(vertical="center", horizontal="center" if col > 7 else "left")
            c.font = _font(size=10)

            if alt:
                c.fill = _fill(_CINZA_CLR)

            # Taxa como %
            if col == 12 and isinstance(val, float):
                c.number_format = "0.0%"

            # Status colorido
            if col == 13:
                c.font = _font(bold=True, color=_status_cor(st), size=10)

            # Faltas em vermelho se > 0
            if col == 10 and val > 0:
                c.font = _font(bold=True, color=_VERMELHO, size=10)

    _auto_width(ws)


# ── Aba 3: Conteúdos ──────────────────────────────────────────────
def _aba_conteudos(wb, dados):
    ws = wb.create_sheet("Conteúdos")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30

    colunas = [
        "Arquivo", "Escola", "Turma", "Disciplina",
        "Trimestre", "Data", "Data Formatada",
        "Tipo de Aula", "Conteúdo / Observações", "Responsável",
    ]
    _header_row(ws, colunas, cor_bg="374151")
    _freeze(ws)

    vistos = set()
    i = 2
    for d in dados:
        chave = (d.get("arquivo", ""), d.get("data", ""))
        if chave in vistos:
            continue
        vistos.add(chave)

        conteudo = d.get("conteudo_obs", "").strip()
        tipo     = d.get("tipo_aula", "").strip()
        if not conteudo and not tipo:
            continue

        alt  = (i % 2 == 0)
        vals = [
            d.get("arquivo", ""), d.get("escola", ""), d.get("turma", ""),
            d.get("disciplina", ""), d.get("trimestre", ""),
            d.get("data", ""), d.get("data_fmt", ""),
            tipo, conteudo, d.get("responsavel", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.border = _border_thin()
            c.alignment = Alignment(vertical="center", wrap_text=(col == 9))
            c.font = _font(size=10)
            if alt:
                c.fill = _fill(_CINZA_CLR)

        ws.row_dimensions[i].height = 40
        i += 1

    _auto_width(ws)
    ws.column_dimensions["I"].width = 70


# ── Entrada principal ─────────────────────────────────────────────
def exportar_xlsx(dados, caminho_saida):
    """
    Gera o Excel com 3 abas a partir da lista de dicts `dados`.
    Cada dict já deve ter os campos calculados (taxa_presenca, total_aulas, data_fmt).
    """
    if not dados:
        return

    wb = Workbook()
    _aba_frequencia(wb, dados)
    _aba_resumo(wb, dados)
    _aba_conteudos(wb, dados)
    wb.save(caminho_saida)
